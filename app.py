"""
app.py  —  Plant Demand Dashboard
Run:  streamlit run app.py
"""
from __future__ import annotations

import base64
import hashlib
import io
import json
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# SUPABASE  —  reads credentials from st.secrets (Streamlit Cloud) or env vars
# ─────────────────────────────────────────────────────────────────────────────

try:
    from supabase import create_client, Client as SupabaseClient
    _SUPA_URL = st.secrets.get("SUPABASE_URL", os.environ.get("SUPABASE_URL", ""))
    _SUPA_KEY = st.secrets.get("SUPABASE_KEY", os.environ.get("SUPABASE_KEY", ""))
    if _SUPA_URL and _SUPA_KEY:
        _supa: SupabaseClient = create_client(_SUPA_URL, _SUPA_KEY)
        SUPABASE_OK = True
    else:
        _supa = None
        SUPABASE_OK = False
except Exception:
    _supa = None
    SUPABASE_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

DATA_DIR = Path("plant_data")
DATA_DIR.mkdir(exist_ok=True)

TIME_STAMPS = []
for i in range(96):
    h_start = i * 15 // 60
    m_start = (i * 15) % 60
    h_end   = (i * 15 + 15) // 60
    m_end   = ((i * 15 + 15)) % 60
    TIME_STAMPS.append(f"{h_start}:{m_start:02d} - {h_end}:{m_end:02d}")

PLANTS = [
    ("GEN_80MW",   "80MW Generation",    "80MW GENERATION",      "80MW AUXILIARY",   "Generation", "MW"),
    ("GEN_43MW",   "43MW Generation",    "43MW GENERATION",      "43MW AUXILIARY",   "Generation", "MW"),
    ("GEN_Solar",  "Solar Generation",   "SOLAR NET GENERATION", "SOLAR AUXILIARY",  "Generation", "MW"),
    ("WLL_WLL",    "WLL",                "WLL",                  None,               "WLL",        "MW"),
    ("WLL_WHSL",   "WHSL",               "WHSL",                 None,               "WLL",        "MW"),
    ("WCL_PIPE",   "WCL Pipe Division",  "WCL PIPE DIVISION",    None,               "WCL",        "MW"),
    ("WCL_STEEL",  "WCL Steel Division", "WCL STEEL DIVISION",   None,               "WCL",        "MW"),
    ("WCL_ATSPL",  "ATSPL",              "ATSPL",                None,               "WCL",        "MW"),
    ("WCL_WDIPL",  "WDIPL",              "WDIPL",                None,               "WCL",        "MW"),
    ("WCL_WASCO",  "WASCO",              "WASCO",                None,               "WCL",        "MW"),
]

PLANT_BY_ID = {p[0]: p for p in PLANTS}

def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

DEFAULT_CREDENTIALS = {p[0]: _hash(p[0].lower()) for p in PLANTS}

# ─────────────────────────────────────────────────────────────────────────────
# LOGO HELPER  —  reads logo file and returns base64 img tag
# ─────────────────────────────────────────────────────────────────────────────

def _logo_html(height: int = 36) -> str:
    """Return an <img> tag with the logo embedded as base64. Falls back gracefully."""
    logo_path = Path("50hertz_Logo (2).png")
    if logo_path.exists():
        data = base64.b64encode(logo_path.read_bytes()).decode()
        return f'<img src="data:image/png;base64,{data}" style="height:{height}px;width:auto;display:block;" />'
    # Fallback: text badge if file not found
    return "<span style='font-size:0.85rem;font-weight:700;color:#1e293b;letter-spacing:0.05em'>50HERTZ</span>"

# ─────────────────────────────────────────────────────────────────────────────
# CREDENTIALS
# ─────────────────────────────────────────────────────────────────────────────

CREDENTIALS_FILE = DATA_DIR / "credentials.json"

def load_credentials() -> dict:
    if SUPABASE_OK:
        try:
            rows = _supa.table("credentials").select("plant_id, password_hash").execute()
            return {r["plant_id"]: r["password_hash"] for r in rows.data}
        except Exception:
            pass
    if not CREDENTIALS_FILE.exists():
        CREDENTIALS_FILE.write_text(json.dumps(DEFAULT_CREDENTIALS, indent=2))
    return json.loads(CREDENTIALS_FILE.read_text())

def verify_password(plant_id: str, password: str) -> bool:
    creds = load_credentials()
    return creds.get(plant_id) == _hash(password)

# ─────────────────────────────────────────────────────────────────────────────
# DATA STORAGE
# ─────────────────────────────────────────────────────────────────────────────

def data_file(plant_id: str) -> Path:
    return DATA_DIR / f"{plant_id}.csv"

def save_data(plant_id: str, df: pd.DataFrame) -> None:
    df = df.copy()
    df["plant_id"] = plant_id

    if SUPABASE_OK:
        try:
            records = []
            for _, row in df.iterrows():
                rec = {
                    "plant_id":   plant_id,
                    "date":       row["Date"].strftime("%Y-%m-%d"),
                    "time_block": int(row["Time Block"]),
                    "value":      None if pd.isna(row["Value"]) else float(row["Value"]),
                    "auxiliary":  None if (pd.isna(row.get("Auxiliary", float("nan"))) if "Auxiliary" in row else True) else float(row["Auxiliary"]),
                    "unit":       str(row.get("Unit", "MW")),
                }
                records.append(rec)
            for i in range(0, len(records), 500):
                _supa.table("plant_readings").upsert(
                    records[i:i+500],
                    on_conflict="plant_id,date,time_block"
                ).execute()
            return
        except Exception as e:
            st.warning(f"Supabase save failed, falling back to local: {e}")

    path = data_file(plant_id)
    if path.exists():
        existing = pd.read_csv(path, parse_dates=["Date"])
        new_dates = df["Date"].unique()
        existing = existing[~existing["Date"].isin(new_dates)]
        combined = pd.concat([existing, df], ignore_index=True)
    else:
        combined = df
    combined.sort_values(["Date", "Time Block"], inplace=True)
    combined.to_csv(path, index=False)


def load_data(plant_id: str) -> Optional[pd.DataFrame]:
    if SUPABASE_OK:
        try:
            rows = _supa.table("plant_readings") \
                        .select("date, time_block, value, auxiliary, unit") \
                        .eq("plant_id", plant_id) \
                        .order("date") \
                        .order("time_block") \
                        .execute()
            if not rows.data:
                return None
            df = pd.DataFrame(rows.data)
            df.rename(columns={"date": "Date", "time_block": "Time Block",
                                "value": "Value", "auxiliary": "Auxiliary",
                                "unit": "Unit"}, inplace=True)
            df["Date"] = pd.to_datetime(df["Date"])
            df["Time Block"] = df["Time Block"].astype(int)
            df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
            df["Auxiliary"] = pd.to_numeric(df["Auxiliary"], errors="coerce")
            df.sort_values(["Date", "Time Block"], inplace=True)
            return df
        except Exception as e:
            st.warning(f"Supabase load failed, falling back to local: {e}")

    path = data_file(plant_id)
    if not path.exists():
        return None
    df = pd.read_csv(path, parse_dates=["Date"])
    df.sort_values(["Date", "Time Block"], inplace=True)
    return df


def load_last_entry(plant_id: str) -> Optional[pd.DataFrame]:
    df = load_data(plant_id)
    if df is None or df.empty:
        return None
    last_date = df["Date"].max()
    return df[df["Date"] == last_date].copy()

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_upload(uploaded_file, plant_id: str) -> tuple[Optional[pd.DataFrame], str]:
    pid, display, col_header, aux_col, group, unit = PLANT_BY_ID[plant_id]

    try:
        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file, header=1)
            cols_upper = [str(c).strip().upper() for c in df.columns]
            if "DATE" not in cols_upper:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, header=0)

        df.columns = [str(c).strip() for c in df.columns]

        date_col = next((c for c in df.columns if "date" in c.lower()), None)
        if date_col is None:
            return None, "Could not find 'Date' column."

        tb_col = next(
            (c for c in df.columns if "time" in c.lower() or "interval" in c.lower() or "block" in c.lower()),
            None
        )
        if tb_col is None:
            return None, "Could not find 'Time Interval' or 'Time Block' column."

        value_col = None
        for c in df.columns:
            if c.upper() == col_header.upper():
                value_col = c
                break
        if value_col is None:
            for c in df.columns:
                if col_header.upper() in c.upper() or c.upper() in col_header.upper():
                    value_col = c
                    break
        if value_col is None:
            return None, f"Could not find column '{col_header}'. Columns: {list(df.columns)}"

        aux_val_col = None
        if aux_col:
            for c in df.columns:
                if c.upper() == aux_col.upper():
                    aux_val_col = c
                    break
            if aux_val_col is None:
                for c in df.columns:
                    if aux_col.upper() in c.upper() or c.upper() in aux_col.upper():
                        aux_val_col = c
                        break

        raw_tb = df[tb_col]
        if pd.api.types.is_numeric_dtype(raw_tb):
            time_block_series = pd.to_numeric(raw_tb, errors="coerce")
        else:
            ts_map = {ts: i+1 for i, ts in enumerate(TIME_STAMPS)}
            ts_map_clean = {ts.replace(" ", ""): i+1 for i, ts in enumerate(TIME_STAMPS)}
            def map_ts(v):
                v = str(v).strip()
                if v in ts_map:
                    return ts_map[v]
                v2 = v.replace(" ", "")
                if v2 in ts_map_clean:
                    return ts_map_clean[v2]
                try:
                    return int(float(v))
                except Exception:
                    return None
            time_block_series = raw_tb.apply(map_ts)

        result = pd.DataFrame({
            "Date":       pd.to_datetime(df[date_col], dayfirst=True, errors="coerce"),
            "Time Block": time_block_series,
            "Value":      pd.to_numeric(df[value_col], errors="coerce"),
            "Unit":       unit,
        })

        if aux_val_col:
            result["Auxiliary"] = pd.to_numeric(df[aux_val_col], errors="coerce")
        else:
            result["Auxiliary"] = None

        result.dropna(subset=["Date", "Time Block", "Value"], inplace=True)

        if result.empty:
            return None, "No valid data rows found after parsing."

        return result, ""

    except Exception as e:
        return None, f"Error reading file: {e}"

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL TEMPLATE GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def _d2_date() -> date:
    return date.today() + timedelta(days=2)

def generate_template(plant_id: str) -> bytes:
    if not OPENPYXL_OK:
        return b""

    pid, display, col_header, aux_col, group, unit = PLANT_BY_ID[plant_id]
    d2 = _d2_date()
    d2_str = d2.strftime("%d-%m-%Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = display[:31]

    HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
    SUBHDR_FILL   = PatternFill("solid", fgColor="2D6A9F")
    ALT_FILL      = PatternFill("solid", fgColor="EBF5FB")
    HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    DATA_FONT     = Font(name="Calibri", size=10)
    CENTER        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="BDC3C7")
    BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = ["Date", "Time Interval", col_header]
    if aux_col:
        cols.append(aux_col)
    num_cols = len(cols)
    merge_end = get_column_letter(num_cols)

    ws.merge_cells(f"A1:{merge_end}1")
    title_cell = ws["A1"]
    title_cell.value = f"{display} — Daily Data Template  |  D+2: {d2_str}"
    title_cell.font  = TITLE_FONT
    title_cell.fill  = HEADER_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    for ci, hdr in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=hdr)
        cell.font      = HEADER_FONT
        cell.fill      = SUBHDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[2].height = 20

    for i, ts in enumerate(TIME_STAMPS):
        row = 3 + i
        fill = ALT_FILL if i % 2 == 0 else None

        dc = ws.cell(row=row, column=1, value=d2_str)
        dc.font = DATA_FONT; dc.alignment = CENTER; dc.border = BORDER
        if fill: dc.fill = fill

        tc = ws.cell(row=row, column=2, value=ts)
        tc.font = DATA_FONT; tc.alignment = CENTER; tc.border = BORDER
        if fill: tc.fill = fill

        vc = ws.cell(row=row, column=3, value=None)
        vc.font = DATA_FONT; vc.alignment = CENTER; vc.border = BORDER
        if fill: vc.fill = fill

        if aux_col:
            ac = ws.cell(row=row, column=4, value=None)
            ac.font = DATA_FONT; ac.alignment = CENTER; ac.border = BORDER
            if fill: ac.fill = fill

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    if aux_col:
        ws.column_dimensions["D"].width = 22

    ws.freeze_panes = "A3"

    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Plant",        display),
        ("Group",        group),
        ("Column",       col_header),
        ("",             ""),
        ("Instructions", ""),
        ("1.",           "Fill in the data values in the highlighted column(s)."),
        ("2.",           f"Date is pre-filled as D+2: {d2_str}. Do not change the Date column."),
        ("3.",           "Time Intervals are fixed 15-min blocks (96 rows = full day)."),
        ("4.",           "Save as .xlsx and upload via the Data Input page."),
        ("5.",           "Do NOT add or remove rows."),
    ]
    for ri, (k, v) in enumerate(instructions, start=1):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=(ri<=4 or k.endswith(".")))
        ws2.cell(row=ri, column=2, value=v)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# CHARTS
# ─────────────────────────────────────────────────────────────────────────────

COLORS = {
    "GEN_80MW":   "#F97316",
    "GEN_43MW":   "#EAB308",
    "GEN_Solar":  "#22C55E",
    "WLL_WLL":    "#3B82F6",
    "WLL_WHSL":   "#6366F1",
    "WCL_PIPE":   "#EC4899",
    "WCL_STEEL":  "#14B8A6",
    "WCL_ATSPL":  "#8B5CF6",
    "WCL_WDIPL":  "#F43F5E",
    "WCL_WASCO":  "#06B6D4",
    "TOTAL_GEN":  "#F97316",
    "TOTAL_WLL":  "#3B82F6",
    "TOTAL_WCL":  "#EC4899",
    "TOTAL_AUX":  "#8B5CF6",
    "TOTAL_LOAD": "#1E293B",
}

def _tb_to_label(tb):
    try:
        idx = int(tb) - 1
        if 0 <= idx < 96:
            return TIME_STAMPS[idx]
    except Exception:
        pass
    return str(tb)

def make_line_chart(df: pd.DataFrame, plant_id: str, title: str) -> go.Figure:
    _, display, _, _, group, unit = PLANT_BY_ID[plant_id]
    color = COLORS.get(plant_id, "#3B82F6")

    fig = go.Figure()

    for d in sorted(df["Date"].unique()):
        sub = df[df["Date"] == d].sort_values("Time Block")
        label = pd.to_datetime(d).strftime("%d %b %Y")
        x_labels = sub["Time Block"].apply(_tb_to_label)
        fig.add_trace(go.Scatter(
            x=x_labels, y=sub["Value"],
            mode="lines", name=label,
            line=dict(width=2),
            hovertemplate=f"%{{x}}<br>%{{y:.2f}} {unit}<br>{label}<extra></extra>",
        ))

        if "Auxiliary" in sub.columns and sub["Auxiliary"].notna().any():
            fig.add_trace(go.Scatter(
                x=x_labels, y=sub["Auxiliary"],
                mode="lines", name=f"{label} (Aux)",
                line=dict(width=1.5, dash="dot", color=color),
                opacity=0.6,
                hovertemplate=f"%{{x}}<br>Aux: %{{y:.2f}} {unit}<br>{label}<extra></extra>",
            ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=16, family="DM Sans", color="#1e293b"), x=0.02),
        xaxis=dict(title="Time Interval", gridcolor="#f1f5f9", linecolor="#e2e8f0",
                   tickfont=dict(family="DM Sans", size=9), tickangle=-45, nticks=12),
        yaxis=dict(title=f"Value ({unit})", gridcolor="#f1f5f9", linecolor="#e2e8f0",
                   tickfont=dict(family="DM Sans", size=11), rangemode="tozero"),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(family="DM Sans", size=11)),
        margin=dict(l=50, r=20, t=60, b=70),
        height=350,
    )
    return fig


def make_total_chart(series_dict: dict, title: str, color_map: dict) -> go.Figure:
    fig = go.Figure()
    for name, ser in series_dict.items():
        if ser is None or ser.empty:
            continue
        x_labels = ser.index.map(_tb_to_label)
        fig.add_trace(go.Scatter(
            x=x_labels, y=ser.values,
            mode="lines", name=name,
            line=dict(width=2.5, color=color_map.get(name, "#888")),
            hovertemplate=f"<b>{name}</b><br>%{{x}}<br>%{{y:.2f}} MW<extra></extra>",
        ))
    fig.update_layout(
        title=dict(text=title, font=dict(size=15, family="DM Sans", color="#1e293b"), x=0.02),
        xaxis=dict(title="Time Interval", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=9), tickangle=-45, nticks=12),
        yaxis=dict(title="MW", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=11), rangemode="tozero"),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        legend=dict(font=dict(family="DM Sans", size=11)),
        margin=dict(l=50, r=20, t=60, b=70),
        height=380,
    )
    return fig


def make_dashboard_overview(all_data: dict) -> go.Figure:
    fig = go.Figure()
    for plant_id, df in all_data.items():
        if df is None or df.empty:
            continue
        last_date = df["Date"].max()
        sub = df[df["Date"] == last_date].sort_values("Time Block")
        _, display, _, _, group, unit = PLANT_BY_ID[plant_id]
        x_labels = sub["Time Block"].apply(_tb_to_label)
        fig.add_trace(go.Scatter(
            x=x_labels, y=sub["Value"],
            mode="lines", name=display,
            line=dict(color=COLORS.get(plant_id, "#888"), width=2),
            hovertemplate=f"<b>{display}</b><br>%{{x}}<br>%{{y:.2f}} {unit}<br>{pd.to_datetime(last_date).strftime('%d %b %Y')}<extra></extra>",
        ))
    fig.update_layout(
        title=dict(text="All Plants — Latest Submission",
                   font=dict(size=16, family="DM Sans", color="#1e293b"), x=0.02),
        xaxis=dict(title="Time Interval", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=9), tickangle=-45, nticks=12),
        yaxis=dict(title="MW", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=11), rangemode="tozero"),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        legend=dict(orientation="v", font=dict(family="DM Sans", size=11),
                    bgcolor="rgba(255,255,255,0.9)", bordercolor="#e2e8f0", borderwidth=1),
        margin=dict(l=50, r=20, t=60, b=70),
        height=420,
    )
    return fig

# ─────────────────────────────────────────────────────────────────────────────
# TOTALS HELPER
# ─────────────────────────────────────────────────────────────────────────────

def compute_totals(plants_with_data: dict, selected_date) -> dict:
    def get_series(plant_id, col="Value"):
        df = plants_with_data.get(plant_id)
        if df is None or df.empty:
            return None
        sub = df[df["Date"] == selected_date]
        if sub.empty:
            sub = df[df["Date"] == df["Date"].max()]
        if sub.empty:
            return None
        if col == "Auxiliary":
            if "Auxiliary" not in sub.columns:
                return None
            s = sub.set_index("Time Block")["Auxiliary"].dropna()
        else:
            s = sub.set_index("Time Block")["Value"]
        return s if not s.empty else None

    idx = pd.RangeIndex(1, 97)

    def safe_add(*series_list):
        result = pd.Series(0.0, index=idx)
        any_data = False
        for s in series_list:
            if s is not None and not s.empty:
                result = result.add(s.reindex(idx, fill_value=0), fill_value=0)
                any_data = True
        return result if any_data else None

    wcl = safe_add(
        get_series("WCL_PIPE"), get_series("WCL_STEEL"),
        get_series("WCL_ATSPL"), get_series("WCL_WDIPL"), get_series("WCL_WASCO"),
    )
    wll = safe_add(get_series("WLL_WLL"), get_series("WLL_WHSL"))
    aux = safe_add(
        get_series("GEN_80MW", "Auxiliary"),
        get_series("GEN_43MW", "Auxiliary"),
        get_series("GEN_Solar", "Auxiliary"),
    )
    total_load = safe_add(
        wcl if wcl is not None else None,
        wll if wll is not None else None,
        aux if aux is not None else None,
    )
    total_gen = safe_add(
        get_series("GEN_80MW"), get_series("GEN_43MW"), get_series("GEN_Solar"),
    )

    return {
        "Total WCL":        wcl,
        "Total WLL":        wll,
        "Total Auxiliary":  aux,
        "Total Load":       total_load,
        "Total Generation": total_gen,
    }

def avg_or_none(s) -> Optional[float]:
    if s is None or (hasattr(s, "empty") and s.empty):
        return None
    return float(s.mean())

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: LOGIN / INPUT
# ─────────────────────────────────────────────────────────────────────────────

def page_input():
    st.markdown("""
    <div style='margin-bottom:1.5rem'>
        <h2 style='font-family:DM Sans,sans-serif;color:#1e293b;font-size:1.6rem;margin:0'>
            📥 Data Input
        </h2>
        <p style='color:#64748b;font-family:DM Sans,sans-serif;margin-top:0.3rem'>
            Log in with your plant credentials and upload your daily demand data.
        </p>
    </div>
    """, unsafe_allow_html=True)

    if "logged_in_plant" not in st.session_state:
        st.session_state.logged_in_plant = None

    if not st.session_state.logged_in_plant:
        with st.container():
            st.markdown("#### 🔐 Plant Login")
            col1, col2 = st.columns([1, 1])
            with col1:
                selected = st.selectbox(
                    "Select Your Plant",
                    options=[p[0] for p in PLANTS],
                    format_func=lambda x: PLANT_BY_ID[x][1],
                )
            with col2:
                password = st.text_input("Password", type="password",
                                          help="Default password is your plant ID in lowercase")

            if st.button("Login →", type="primary"):
                if verify_password(selected, password):
                    st.session_state.logged_in_plant = selected
                    st.success(f"Welcome, {PLANT_BY_ID[selected][1]}!")
                    st.rerun()
                else:
                    st.error("Incorrect password.")

            st.markdown("""
            <div style='background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
                        padding:0.8rem 1rem;margin-top:1rem;font-family:DM Sans,sans-serif;
                        font-size:0.85rem;color:#0369a1'>
                💡 <strong>Default passwords</strong> are your Plant ID in lowercase.<br>
                Examples: <code>gen_80mw</code>, <code>wll_wll</code>, <code>wcl_pipe</code>
            </div>
            """, unsafe_allow_html=True)

        if OPENPYXL_OK:
            st.markdown("---")
            d2 = _d2_date()
            sel_pid, sel_disp, sel_col, sel_aux, sel_grp, sel_unit = PLANT_BY_ID[selected]
            aux_note = f" + {sel_aux}" if sel_aux else ""
            st.markdown(f"""
            <div style='background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;
                        padding:0.7rem 1rem;margin-bottom:0.8rem;font-family:DM Sans,sans-serif;
                        font-size:0.85rem;color:#166534'>
                📋 <strong>{sel_disp}</strong> template — columns: Date | Time Interval |
                <strong>{sel_col}</strong>{aux_note}<br>
                Pre-filled for <strong>D+2: {d2.strftime('%d %B %Y')}</strong> · 96 rows
            </div>
            """, unsafe_allow_html=True)
            tmpl_bytes = generate_template(selected)
            st.download_button(
                label=f"⬇️ Download Template — {sel_disp}",
                data=tmpl_bytes,
                file_name=f"template_{selected}_{d2.strftime('%d-%m-%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="prelogin_tmpl",
                use_container_width=False,
            )
        return

    plant_id = st.session_state.logged_in_plant
    pid, display, col_header, aux_col, group, unit = PLANT_BY_ID[plant_id]

    col_title, col_logout = st.columns([4, 1])
    with col_title:
        color = COLORS.get(plant_id, "#3B82F6")
        st.markdown(f"""
        <div style='display:flex;align-items:center;gap:0.6rem;margin-bottom:1rem'>
            <div style='width:12px;height:12px;border-radius:50%;background:{color}'></div>
            <span style='font-family:DM Sans,sans-serif;font-size:1.1rem;font-weight:600;color:#1e293b'>
                Logged in as: {display}
            </span>
            <span style='background:#f1f5f9;color:#64748b;font-size:0.75rem;padding:2px 8px;
                         border-radius:12px;font-family:DM Sans,sans-serif'>{group}</span>
        </div>
        """, unsafe_allow_html=True)
    with col_logout:
        if st.button("Logout", use_container_width=True):
            st.session_state.logged_in_plant = None
            st.rerun()

    st.markdown("#### 📥 Download Data Template")
    d2_str = _d2_date().strftime("%d-%m-%Y")
    aux_note = f" + {aux_col}" if aux_col else ""
    st.markdown(f"""
    <div style='background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;
                padding:0.8rem 1rem;margin-bottom:0.8rem;font-family:DM Sans,sans-serif;
                font-size:0.85rem;color:#166534'>
        📋 Template columns: <strong>Date</strong> | <strong>Time Interval</strong> | 
        <strong>{col_header}</strong>{aux_note}<br>
        Date pre-filled as <strong>D+2: {d2_str}</strong> · 96 rows (00:00–24:00)
    </div>
    """, unsafe_allow_html=True)

    if OPENPYXL_OK:
        tmpl_bytes = generate_template(plant_id)
        st.download_button(
            label=f"⬇️  Download Template — {display}",
            data=tmpl_bytes,
            file_name=f"template_{plant_id}_{d2_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
    else:
        st.warning("openpyxl not installed — template download unavailable.")

    last = load_last_entry(plant_id)
    if last is not None:
        last_date = last["Date"].max().strftime("%d %B %Y")
        avg_val = last["Value"].mean()
        max_val = last["Value"].max()
        aux_str = ""
        if "Auxiliary" in last.columns and last["Auxiliary"].notna().any():
            avg_aux = last["Auxiliary"].mean()
            aux_str = f"<div style='color:#64748b;font-size:0.9rem'>Avg Aux: <strong>{avg_aux:.2f} {unit}</strong></div>"
        st.markdown(f"""
        <div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;
                    padding:1rem 1.2rem;margin:1rem 0;font-family:DM Sans,sans-serif'>
            <div style='font-size:0.8rem;color:#94a3b8;margin-bottom:0.3rem'>LAST SUBMISSION</div>
            <div style='display:flex;gap:2rem;align-items:center;flex-wrap:wrap'>
                <div><span style='font-size:1rem;font-weight:600;color:#1e293b'>{last_date}</span></div>
                <div style='color:#64748b;font-size:0.9rem'>Avg: <strong>{avg_val:.2f} {unit}</strong></div>
                <div style='color:#64748b;font-size:0.9rem'>Peak: <strong>{max_val:.2f} {unit}</strong></div>
                <div style='color:#64748b;font-size:0.9rem'>Rows: <strong>{len(last)}</strong></div>
                {aux_str}
            </div>
        </div>
        """, unsafe_allow_html=True)

        with st.expander("Preview last submission data"):
            show = last[["Date", "Time Block", "Value"]].copy()
            show["Time Interval"] = show["Time Block"].apply(_tb_to_label)
            show["Date"] = show["Date"].dt.strftime("%d-%m-%Y")
            cols_show = ["Date", "Time Interval", col_header]
            show = show.rename(columns={"Value": col_header})
            if "Auxiliary" in last.columns and last["Auxiliary"].notna().any():
                show[aux_col] = last["Auxiliary"].values
                cols_show.append(aux_col)
            st.dataframe(show[cols_show], use_container_width=True, hide_index=True)
    else:
        st.info("No previous data found. Download the template, fill in your values, and upload below.")

    st.markdown("#### 📂 Upload Today's Data")
    aux_note2 = f", <strong>{aux_col}</strong>" if aux_col else ""
    st.markdown(f"""
    <div style='background:#fefce8;border:1px solid #fde68a;border-radius:8px;
                padding:0.8rem 1rem;margin-bottom:1rem;font-family:DM Sans,sans-serif;
                font-size:0.85rem;color:#92400e'>
        📋 Required columns: <strong>Date</strong>, <strong>Time Interval</strong>, 
        <strong>{col_header}</strong>{aux_note2}<br>
        Accepted: <strong>.xlsx</strong> or <strong>.csv</strong> &nbsp;|&nbsp; 96 rows
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader(
        f"Upload file for {display}",
        type=["xlsx", "xls", "csv"],
    )

    if uploaded:
        df_parsed, err = parse_upload(uploaded, plant_id)
        if err:
            st.error(f"❌ {err}")
        else:
            dates_found = df_parsed["Date"].dt.strftime("%d %B %Y").unique()
            st.success(f"✅ Parsed {len(df_parsed)} rows for: {', '.join(dates_found)}")

            preview = df_parsed[["Date", "Time Block", "Value"]].copy()
            preview["Time Interval"] = preview["Time Block"].apply(_tb_to_label)
            preview["Date"] = preview["Date"].dt.strftime("%d-%m-%Y")
            preview = preview.rename(columns={"Value": col_header})
            show_cols = ["Date", "Time Interval", col_header]
            if "Auxiliary" in df_parsed.columns and df_parsed["Auxiliary"].notna().any():
                preview[aux_col] = df_parsed["Auxiliary"].values
                show_cols.append(aux_col)

            with st.expander("Preview parsed data", expanded=True):
                st.dataframe(preview[show_cols].head(20), use_container_width=True, hide_index=True)
                if len(preview) > 20:
                    st.caption(f"... and {len(preview) - 20} more rows")

            col_a, col_b = st.columns([1, 3])
            with col_a:
                if st.button("✅ Confirm & Save", type="primary", use_container_width=True):
                    save_data(plant_id, df_parsed)
                    st.success("Data saved successfully!")
                    st.balloons()
                    st.rerun()

    if last is not None and len(last) > 0:
        st.markdown("#### 📈 Your Latest Data")
        fig = make_line_chart(last, plant_id, f"{display} — Last Submission")
        st.plotly_chart(fig, use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def page_dashboard():
    st.markdown("""
    <div style='margin-bottom:1.5rem'>
        <h2 style='font-family:DM Sans,sans-serif;color:#1e293b;font-size:1.6rem;margin:0'>
            📊 Demand Dashboard
        </h2>
        <p style='color:#64748b;font-family:DM Sans,sans-serif;margin-top:0.3rem'>
            Real-time view of all plant demand submissions.
        </p>
    </div>
    """, unsafe_allow_html=True)

    all_data = {p[0]: load_data(p[0]) for p in PLANTS}
    plants_with_data = {k: v for k, v in all_data.items() if v is not None and not v.empty}

    if not plants_with_data:
        st.warning("No data available yet. Plants need to submit their data first.")
        return

    all_dates = sorted(set(
        d for df in plants_with_data.values() for d in df["Date"].unique()
    ), reverse=True)
    date_options_str = [pd.to_datetime(d).strftime("%d %B %Y") for d in all_dates]
    selected_date_str = st.selectbox("📅 View Date", date_options_str, key="dash_date")
    selected_date = pd.to_datetime(selected_date_str, format="%d %B %Y")

    st.markdown("#### Summary — Individual Plants")
    card_cols = st.columns(5)
    for i, (plant_id, df) in enumerate(plants_with_data.items()):
        _, display, _, _, group, unit = PLANT_BY_ID[plant_id]
        sub = df[df["Date"] == selected_date]
        if sub.empty:
            sub = df[df["Date"] == df["Date"].max()]
        avg = sub["Value"].mean()
        color = COLORS.get(plant_id, "#3B82F6")
        with card_cols[i % 5]:
            st.markdown(f"""
            <div style='background:white;border:1px solid #e2e8f0;border-top:3px solid {color};
                        border-radius:10px;padding:0.9rem;margin-bottom:0.8rem;font-family:DM Sans,sans-serif'>
                <div style='font-size:0.72rem;color:#94a3b8;text-transform:uppercase;letter-spacing:0.05em'>{group}</div>
                <div style='font-size:0.95rem;font-weight:700;color:#1e293b;margin:0.2rem 0'>{display}</div>
                <div style='font-size:1.3rem;font-weight:800;color:{color}'>{avg:.1f}</div>
                <div style='font-size:0.75rem;color:#94a3b8'>{unit} avg</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 📊 Totals Overview")

    totals = compute_totals(plants_with_data, selected_date)

    TOTAL_META = {
        "Total WCL":        ("WCL Load", "#EC4899",  "Pipe + Steel + ATSPL + WDIPL + WASCO"),
        "Total WLL":        ("WLL Load",  "#3B82F6",  "WLL + WHSL"),
        "Total Auxiliary":  ("Auxiliary", "#8B5CF6",  "80MW Aux + 43MW Aux + Solar Aux"),
        "Total Load":       ("Net Load",  "#1E293B",  "WCL + WLL + Auxiliary"),
        "Total Generation": ("Generation","#F97316",  "80MW + 43MW + Solar"),
    }

    total_cols = st.columns(5)
    for i, (key, (label, color, tooltip)) in enumerate(TOTAL_META.items()):
        s = totals.get(key)
        avg = avg_or_none(s)
        val_str = f"{avg:.1f}" if avg is not None else "—"
        with total_cols[i]:
            st.markdown(f"""
            <div style='background:white;border:2px solid {color};border-radius:10px;
                        padding:0.9rem;margin-bottom:0.8rem;font-family:DM Sans,sans-serif;
                        box-shadow:0 2px 8px rgba(0,0,0,0.06)'>
                <div style='font-size:0.72rem;color:#94a3b8;text-transform:uppercase;
                             letter-spacing:0.05em'>{label}</div>
                <div style='font-size:0.9rem;font-weight:700;color:#1e293b;margin:0.2rem 0'>{key}</div>
                <div style='font-size:1.5rem;font-weight:800;color:{color}'>{val_str}</div>
                <div style='font-size:0.7rem;color:#94a3b8;margin-top:2px'>{tooltip}</div>
            </div>
            """, unsafe_allow_html=True)

    total_color_map = {
        "Total WCL":        "#EC4899",
        "Total WLL":        "#3B82F6",
        "Total Auxiliary":  "#8B5CF6",
        "Total Load":       "#1E293B",
        "Total Generation": "#F97316",
    }
    valid_totals = {k: v for k, v in totals.items() if v is not None}
    if valid_totals:
        fig_totals = make_total_chart(valid_totals, "Totals — All Groups", total_color_map)
        st.plotly_chart(fig_totals, use_container_width=True)

    st.markdown("---")
    st.markdown("#### All Plants Overview")
    fig_overview = make_dashboard_overview(plants_with_data)
    st.plotly_chart(fig_overview, use_container_width=True)

    st.markdown("---")
    st.markdown("#### Detailed Plant Charts")

    group_filter = st.radio(
        "Plant Group", ["All", "Generation", "WLL", "WCL"],
        horizontal=True,
    )

    filtered_plants = [
        (pid, df) for pid, df in plants_with_data.items()
        if group_filter == "All" or PLANT_BY_ID[pid][4] == group_filter
    ]

    if not filtered_plants:
        st.info("No data for selected filters.")
    else:
        for i in range(0, len(filtered_plants), 2):
            row_cols = st.columns(2)
            for j, (plant_id, df) in enumerate(filtered_plants[i:i+2]):
                _, display, _, _, group, unit = PLANT_BY_ID[plant_id]
                with row_cols[j]:
                    date_df = df[df["Date"] == selected_date]
                    is_fallback = False
                    display_date = selected_date
                    if date_df.empty:
                        last_available = df["Date"].max()
                        if pd.notnull(last_available):
                            date_df = df[df["Date"] == last_available]
                            display_date = last_available
                            is_fallback = True

                    if date_df.empty:
                        st.markdown(f"""
                        <div style='background:#f8fafc;border:1px dashed #cbd5e1;border-radius:10px;
                                    padding:2rem;text-align:center;color:#94a3b8'>
                            <strong>{display}</strong><br>No data available
                        </div>""", unsafe_allow_html=True)
                    else:
                        if is_fallback:
                            st.markdown(f"""
                            <div style='background:#fffbeb;border:1px solid #fef3c7;border-radius:6px;
                                        padding:4px 10px;margin-bottom:8px;font-size:0.75rem;color:#92400e'>
                                ⚠️ Showing latest data from <b>{pd.to_datetime(display_date).strftime('%d %b %Y')}</b>
                            </div>""", unsafe_allow_html=True)
                        fig = make_line_chart(date_df, plant_id, display)
                        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.markdown("#### Daily Statistics Table")

    rows = []
    for plant_id, df in plants_with_data.items():
        _, display, _, _, group, unit = PLANT_BY_ID[plant_id]
        for d in sorted(df["Date"].unique(), reverse=True)[:7]:
            sub = df[df["Date"] == d]
            row = {
                "Plant": display, "Group": group,
                "Date":  pd.to_datetime(d).strftime("%d %b %Y"),
                "Rows":  len(sub),
                f"Avg ({unit})": round(sub["Value"].mean(), 2),
                f"Max ({unit})": round(sub["Value"].max(), 2),
                f"Min ({unit})": round(sub["Value"].min(), 2),
            }
            if "Auxiliary" in sub.columns and sub["Auxiliary"].notna().any():
                row["Avg Aux (MW)"] = round(sub["Auxiliary"].mean(), 2)
            rows.append(row)

    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("#### Compare Plants Side-by-Side")

    compare_options = [p[0] for p in PLANTS if p[0] in plants_with_data]
    selected_compare = st.multiselect(
        "Select plants to compare",
        options=compare_options,
        default=compare_options[:3],
        format_func=lambda x: PLANT_BY_ID[x][1],
    )

    if selected_compare:
        fig_compare = go.Figure()
        for pid in selected_compare:
            df = plants_with_data[pid]
            sub = df[df["Date"] == selected_date]
            if sub.empty:
                sub = df[df["Date"] == df["Date"].max()]
            sub = sub.sort_values("Time Block")
            _, display, _, _, _, unit = PLANT_BY_ID[pid]
            x_labels = sub["Time Block"].apply(_tb_to_label)
            fig_compare.add_trace(go.Scatter(
                x=x_labels, y=sub["Value"],
                mode="lines", name=display,
                line=dict(color=COLORS.get(pid, "#888"), width=2.5),
                hovertemplate=f"{display}: %{{y:.2f}} {unit}<extra></extra>",
            ))
        fig_compare.update_layout(
            title=dict(text="Plant Comparison",
                       font=dict(size=15, family="DM Sans", color="#1e293b"), x=0.02),
            xaxis=dict(title="Time Interval", gridcolor="#f1f5f9",
                       tickfont=dict(family="DM Sans", size=9), tickangle=-45, nticks=12),
            yaxis=dict(title="MW", gridcolor="#f1f5f9",
                       tickfont=dict(family="DM Sans", size=11), rangemode="tozero"),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            legend=dict(font=dict(family="DM Sans", size=11)),
            margin=dict(l=50, r=20, t=60, b=70), height=400,
        )
        st.plotly_chart(fig_compare, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: CONSOLIDATED VIEW
# ─────────────────────────────────────────────────────────────────────────────

def page_consolidated():
    st.markdown("""
    <div style='margin-bottom:1.5rem'>
        <h2 style='font-family:DM Sans,sans-serif;color:#1e293b;font-size:1.6rem;margin:0'>
            📑 Consolidated Data
        </h2>
        <p style='color:#64748b;font-family:DM Sans,sans-serif;margin-top:0.3rem'>
            All plants combined into one table — latest available data used where a date is missing.
        </p>
    </div>
    """, unsafe_allow_html=True)

    all_data = {p[0]: load_data(p[0]) for p in PLANTS}
    plants_with_data = {k: v for k, v in all_data.items() if v is not None and not v.empty}

    if not plants_with_data:
        st.warning("No data available yet.")
        return

    all_dates = sorted(set(
        d for df in plants_with_data.values() for d in df["Date"].unique()
    ), reverse=True)
    date_options_str = [pd.to_datetime(d).strftime("%d %B %Y") for d in all_dates]
    selected_date_str = st.selectbox("📅 View Date", date_options_str, key="cons_date")
    selected_date = pd.to_datetime(selected_date_str, format="%d %B %Y")

    def get_plant_series(plant_id, col="Value") -> pd.Series:
        df = plants_with_data.get(plant_id)
        if df is None or df.empty:
            return pd.Series(dtype=float)
        sub = df[df["Date"] == selected_date]
        if sub.empty:
            sub = df[df["Date"] == df["Date"].max()]
        if sub.empty:
            return pd.Series(dtype=float)
        if col == "Auxiliary":
            if "Auxiliary" not in sub.columns:
                return pd.Series(dtype=float)
            return sub.set_index("Time Block")["Auxiliary"]
        return sub.set_index("Time Block")["Value"]

    idx = list(range(1, 97))

    rows = []
    for tb in idx:
        def v(pid, col="Value"):
            s = get_plant_series(pid, col)
            return round(s.get(tb, float("nan")), 3) if not s.empty else float("nan")

        gen_80   = v("GEN_80MW")
        gen_43   = v("GEN_43MW")
        gen_sol  = v("GEN_Solar")
        aux_80   = v("GEN_80MW", "Auxiliary")
        aux_43   = v("GEN_43MW", "Auxiliary")
        aux_sol  = v("GEN_Solar", "Auxiliary")

        total_gen = sum(x for x in [gen_80, gen_43, gen_sol] if pd.notna(x)) or float("nan")
        total_aux = sum(x for x in [aux_80, aux_43, aux_sol] if pd.notna(x)) or float("nan")

        wll  = v("WLL_WLL")
        whsl = v("WLL_WHSL")
        pipe = v("WCL_PIPE")
        stl  = v("WCL_STEEL")
        atspl= v("WCL_ATSPL")
        wdipl= v("WCL_WDIPL")
        wasco= v("WCL_WASCO")

        wml_vals = [x for x in [wll, whsl] if pd.notna(x)]
        wml = sum(wml_vals) if wml_vals else float("nan")

        wcl_vals = [x for x in [pipe, stl, atspl, wdipl, wasco] if pd.notna(x)]
        total_wcl = sum(wcl_vals) if wcl_vals else float("nan")

        demand_vals = [x for x in [total_wcl, wml, total_aux] if pd.notna(x)]
        total_demand = sum(demand_vals) if demand_vals else float("nan")

        rows.append({
            "Time Interval":        TIME_STAMPS[tb - 1],
            "80MW GENERATION":      gen_80,
            "43MW GENERATION":      gen_43,
            "SOLAR NET GENERATION": gen_sol,
            "TOTAL TG GENERATION":  round(total_gen, 3) if pd.notna(total_gen) else float("nan"),
            "80MW AUXILIARY":       aux_80,
            "43MW AUXILIARY":       aux_43,
            "SOLAR AUXILIARY":      aux_sol,
            "TOTAL AUXILIARY":      round(total_aux, 3) if pd.notna(total_aux) else float("nan"),
            "WLL":                  wll,
            "WHSL":                 whsl,
            "WCL PIPE DIVISION":    pipe,
            "WCL STEEL DIVISION":   stl,
            "ATSPL":                atspl,
            "WDIPL":                wdipl,
            "WASCO":                wasco,
            "WML":                  round(wml, 3) if pd.notna(wml) else float("nan"),
            "TOTAL DEMAND":         round(total_demand, 3) if pd.notna(total_demand) else float("nan"),
        })

    cons_df = pd.DataFrame(rows)

    fallback_notes = []
    for pid, df in plants_with_data.items():
        _, disp, _, _, _, _ = PLANT_BY_ID[pid]
        if selected_date not in df["Date"].values:
            fb_date = pd.to_datetime(df["Date"].max()).strftime("%d %b %Y")
            fallback_notes.append(f"<b>{disp}</b>: using {fb_date}")

    if fallback_notes:
        st.markdown(f"""
        <div style='background:#fffbeb;border:1px solid #fde68a;border-radius:8px;
                    padding:0.7rem 1rem;margin-bottom:1rem;font-family:DM Sans,sans-serif;
                    font-size:0.82rem;color:#92400e'>
            ⚠️ <strong>Fallback data used for:</strong> {" &nbsp;|&nbsp; ".join(fallback_notes)}
        </div>
        """, unsafe_allow_html=True)

    numeric_cols = [c for c in cons_df.columns if c != "Time Interval"]
    summary = cons_df[numeric_cols].mean().round(2)
    scols = st.columns(3)
    highlights = [
        ("TOTAL AUXILIARY",    "#8B5CF6"),
        ("TOTAL TG GENERATION","#F97316"),
        ("TOTAL DEMAND",       "#1E293B"),
    ]
    for i, (col, color) in enumerate(highlights):
        val = summary.get(col, float("nan"))
        val_str = f"{val:.2f}" if pd.notna(val) else "—"
        with scols[i]:
            st.markdown(f"""
            <div style='background:white;border:2px solid {color};border-radius:10px;
                        padding:1rem 1.2rem;font-family:DM Sans,sans-serif;
                        box-shadow:0 2px 8px rgba(0,0,0,0.06)'>
                <div style='font-size:0.72rem;color:#94a3b8;text-transform:uppercase;
                             letter-spacing:0.06em;margin-bottom:0.2rem'>{col}</div>
                <div style='font-size:1.8rem;font-weight:800;color:{color}'>{val_str}</div>
                <div style='font-size:0.75rem;color:#94a3b8'>MW avg across 96 blocks</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown(f"##### All 96 Time Blocks — {selected_date_str}")
    st.dataframe(
        cons_df,
        use_container_width=True,
        hide_index=True,
        height=500,
        column_config={c: st.column_config.NumberColumn(format="%.3f") for c in numeric_cols},
    )

    if OPENPYXL_OK:
        def to_excel(df: pd.DataFrame) -> bytes:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Consolidated")
                ws = writer.sheets["Consolidated"]
                hdr_fill = PatternFill("solid", fgColor="1E3A5F")
                hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                for col_cells in ws.columns:
                    max_len = max(len(str(c.value or "")) for c in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = max(12, max_len + 2)
                ws.freeze_panes = "B2"
            return buf.getvalue()

        xl_bytes = to_excel(cons_df)
        d_str = selected_date.strftime("%d-%m-%Y")
        st.download_button(
            label="⬇️ Download Consolidated Excel",
            data=xl_bytes,
            file_name=f"consolidated_{d_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")
    st.markdown("##### Totals Chart")
    chart_cols = {
        "TOTAL TG GENERATION": "#F97316",
        "TOTAL AUXILIARY":     "#8B5CF6",
        "TOTAL DEMAND":        "#1E293B",
    }
    fig = go.Figure()
    for col, color in chart_cols.items():
        series = cons_df[col].dropna()
        if series.empty:
            continue
        fig.add_trace(go.Scatter(
            x=cons_df.loc[series.index, "Time Interval"],
            y=series.values,
            mode="lines", name=col,
            line=dict(color=color, width=2.5),
            hovertemplate=f"<b>{col}</b><br>%{{x}}<br>%{{y:.3f}} MW<extra></extra>",
        ))
    fig.update_layout(
        xaxis=dict(title="Time Interval", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=9), tickangle=-45, nticks=12),
        yaxis=dict(title="MW", gridcolor="#f1f5f9", rangemode="tozero"),
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        legend=dict(font=dict(family="DM Sans", size=11)),
        margin=dict(l=50, r=20, t=40, b=70), height=380,
    )
    st.plotly_chart(fig, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# PAGE: LIVE GENERATION vs DEMAND
# ─────────────────────────────────────────────────────────────────────────────

def page_live():
    st.markdown("""
    <div style='margin-bottom:1rem'>
        <h2 style='font-family:DM Sans,sans-serif;color:#1e293b;font-size:1.6rem;margin:0'>
            ⚡ Live Generation vs Demand
        </h2>
        <p style='color:#64748b;font-family:DM Sans,sans-serif;margin-top:0.3rem'>
            Welspun Power Market — RE Schedule vs Actual Demand &amp; Market Trading
        </p>
    </div>
    """, unsafe_allow_html=True)

    import math as _math_mkt
    import random as _rnd

    slots = 96
    re_schedule   = []
    re_generation  = []
    actual_demand  = []

    _rnd.seed(42)

    for i in range(slots):
        h = (i * 15) / 60.0
        solar   = 30 * _math_mkt.exp(-((h - 12) / 4) ** 2) if 6 <= h <= 18 else 0
        thermal = (62
                   + 8  * _math_mkt.sin(_math_mkt.pi * (h - 2) / 12)
                   + 5  * _math_mkt.sin(_math_mkt.pi * (h - 8) / 6))
        sched   = thermal + solar
        re_schedule.append(round(sched, 2))

        shortfall = (0.10
                     + 0.05 * _math_mkt.sin(2 * _math_mkt.pi * h / 7   + 0.5)
                     + 0.02 * _math_mkt.sin(2 * _math_mkt.pi * h / 1.8 + 1.1))
        re_gen = sched * (1 - shortfall) + (_rnd.random() - 0.5) * 1.5
        re_generation.append(round(max(0, re_gen), 2))

        demand = (sched
                  + 20 * _math_mkt.sin(2 * _math_mkt.pi * h / 5.2 + 0.9)
                  + 12 * _math_mkt.sin(2 * _math_mkt.pi * h / 2.6 + 1.5)
                  + 7  * _math_mkt.sin(2 * _math_mkt.pi * h / 1.4 + 0.4)
                  + (_rnd.random() - 0.5) * 5)
        demand = max(50, min(180, demand))
        actual_demand.append(round(demand, 2))

    net_demand = [round(actual_demand[i] - re_generation[i], 2) for i in range(slots)]
    buy_vals   = [v if v > 0 else 0 for v in net_demand]
    sell_vals  = [-v if v < 0 else 0 for v in net_demand]

    time_labels = []
    for i in range(slots):
        h = i * 15 // 60
        m = (i * 15) % 60
        time_labels.append(f"{h}:{m:02d}")

    crossings = sum(1 for i in range(1, slots) if (net_demand[i-1] > 0) != (net_demand[i] > 0))

    avg_demand  = round(sum(actual_demand)  / slots, 1)
    avg_sched   = round(sum(re_schedule)    / slots, 1)
    avg_gen     = round(sum(re_generation)  / slots, 1)
    total_buy   = round(sum(buy_vals)  * 0.25, 1)
    total_sell  = round(sum(sell_vals) * 0.25, 1)

    st.markdown("#### 📈 Welspun Power Market — RE Schedule vs Actual Demand & Market Trading")
    st.markdown(
        "<div style='font-size:0.82rem;color:#64748b;margin-bottom:0.8rem'>"
        "Manikaran facilitates buy/sell of net demand in open power exchange market on behalf of Welspun. "
        "Green bars = surplus (sell to market) · Red bars = deficit (buy from market).</div>",
        unsafe_allow_html=True,
    )

    mk1, mk2, mk3, mk4, mk5 = st.columns(5)
    for col, label, val, color in [
        (mk1, "Avg Actual Demand",  f"{avg_demand} MW",  "#8b3dba"),
        (mk2, "Avg RE Schedule",    f"{avg_sched} MW",   "#2a5fa5"),
        (mk3, "Avg RE Generation",  f"{avg_gen} MW",     "#1a8a50"),
        (mk4, "Total Market Buy",   f"{total_buy} MWh",  "#b94a2a"),
        (mk5, "Total Market Sell",  f"{total_sell} MWh", "#1a6b44"),
    ]:
        with col:
            st.markdown(f"""
            <div style='background:white;border:1px solid #e2e8f0;border-radius:8px;
                        padding:0.75rem 1rem;font-family:DM Sans,sans-serif'>
                <div style='font-size:0.68rem;color:#94a3b8;text-transform:uppercase;
                            letter-spacing:0.05em;margin-bottom:3px'>{label}</div>
                <div style='font-size:1.25rem;font-weight:700;color:{color}'>{val}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown(
        f"<div style='font-size:0.78rem;color:#64748b;margin:0.5rem 0'>"
        f"⚡ {crossings} crossover events today</div>",
        unsafe_allow_html=True,
    )

    fig_mkt = go.Figure()
    fig_mkt.add_trace(go.Scatter(
        x=time_labels, y=re_schedule,
        mode="lines", line=dict(width=0),
        fill=None, showlegend=False, hoverinfo="skip",
    ))
    fig_mkt.add_trace(go.Scatter(
        x=time_labels, y=re_generation,
        mode="lines", line=dict(width=0),
        fill="tonexty", fillcolor="rgba(42,95,165,0.07)",
        showlegend=False, hoverinfo="skip",
    ))
    fig_mkt.add_trace(go.Scatter(
        x=time_labels, y=re_schedule,
        mode="lines", name="RE Schedule",
        line=dict(color="#2a5fa5", width=2.5),
        hovertemplate="<b>RE Schedule</b><br>%{x}<br>%{y:.1f} MW<extra></extra>",
    ))
    fig_mkt.add_trace(go.Scatter(
        x=time_labels, y=re_generation,
        mode="lines", name="RE Generation (actual)",
        line=dict(color="#1a8a50", width=2, dash="dash"),
        hovertemplate="<b>RE Generation</b><br>%{x}<br>%{y:.1f} MW<extra></extra>",
    ))
    fig_mkt.add_trace(go.Scatter(
        x=time_labels, y=actual_demand,
        mode="lines", name="Actual Demand",
        line=dict(color="#8b3dba", width=2.5, dash="dot"),
        hovertemplate="<b>Actual Demand</b><br>%{x}<br>%{y:.1f} MW<extra></extra>",
    ))
    fig_mkt.update_layout(
        height=380,
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        margin=dict(l=50, r=20, t=20, b=40),
        legend=dict(font=dict(family="DM Sans", size=11),
                    orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1),
        xaxis=dict(gridcolor="#f1f5f9", tickfont=dict(family="DM Sans", size=10),
                   tickangle=-45, tickmode="array",
                   tickvals=time_labels[::8], ticktext=time_labels[::8]),
        yaxis=dict(title="Power (MW)", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=10),
                   ticksuffix=" MW", range=[30, None]),
        hovermode="x unified",
    )
    st.plotly_chart(fig_mkt, use_container_width=True)

    fig_net = go.Figure()
    fig_net.add_trace(go.Bar(
        x=time_labels, y=sell_vals,
        name="Sell to market",
        marker_color="rgba(20,120,70,0.75)",
        hovertemplate="<b>Sell to market</b><br>%{x}<br>%{y:.1f} MW<extra></extra>",
    ))
    fig_net.add_trace(go.Bar(
        x=time_labels, y=[-v for v in buy_vals],
        name="Buy from market",
        marker_color="rgba(180,60,30,0.75)",
        hovertemplate="<b>Buy from market</b><br>%{x}<br>%{customdata:.1f} MW<extra></extra>",
        customdata=buy_vals,
    ))
    fig_net.update_layout(
        height=180, barmode="relative",
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        margin=dict(l=50, r=20, t=10, b=40),
        legend=dict(font=dict(family="DM Sans", size=11),
                    orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        xaxis=dict(gridcolor="#f1f5f9", tickfont=dict(family="DM Sans", size=10),
                   tickangle=-45, tickmode="array",
                   tickvals=time_labels[::8], ticktext=time_labels[::8]),
        yaxis=dict(title="Net demand", gridcolor="#f1f5f9",
                   tickfont=dict(family="DM Sans", size=10), ticksuffix=" MW"),
        hovermode="x unified",
    )
    st.plotly_chart(fig_net, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Plant Demand Dashboard",
        page_icon="⚡",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── Global styles ─────────────────────────────────────────────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700;800&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .main .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1200px; }
    .stButton > button { font-family: 'DM Sans', sans-serif; font-weight: 600; border-radius: 8px; }
    .stButton > button[kind="primary"] { background: #1e40af; border: none; }
    .logo-top-right {
        position: fixed;
        top: 0.45rem;
        right: 4.5rem;
        z-index: 9999;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Logo: base64-encoded so no static folder needed ───────────────────
    st.markdown(
        f'<div class="logo-top-right">{_logo_html(height=36)}</div>',
        unsafe_allow_html=True,
    )

    # ── Sidebar ───────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("""
        <div style='padding:0.5rem 0 1.5rem 0'>
            <div style='font-size:1.5rem;font-weight:800;color:#1e293b'>⚡ Welspun</div>
            <div style='font-size:0.8rem;color:#94a3b8'>WLL / WCL Automation</div>
        </div>
        """, unsafe_allow_html=True)

        page = st.radio(
            "Navigation",
            ["📥 Data Input", "📊 Dashboard", "📑 Consolidated", "⚡ Live Generation"],
            label_visibility="collapsed",
        )

        st.markdown("---")
        st.markdown(
            "<div style='font-size:0.75rem;color:#94a3b8;font-weight:600;"
            "text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.5rem'>"
            "Plant Status</div>",
            unsafe_allow_html=True,
        )

        for p in PLANTS:
            pid, display, _, _, group, _ = p
            df = load_data(pid)
            if df is not None and not df.empty:
                last_date = df["Date"].max()
                is_today = pd.to_datetime(last_date).date() >= date.today()
                dot = "🟢" if is_today else "🟡"
                date_str = pd.to_datetime(last_date).strftime("%d %b")
            else:
                dot = "🔴"
                date_str = "No data"
            st.markdown(f"""
            <div style='display:flex;justify-content:space-between;align-items:center;
                        padding:0.25rem 0;font-size:0.82rem'>
                <span style='color:#334155'>{dot} {display}</span>
                <span style='color:#94a3b8;font-size:0.75rem'>{date_str}</span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")
        d2 = _d2_date()
        st.markdown(f"""
        <div style='font-size:0.75rem;color:#94a3b8'>
            Last refreshed<br>{datetime.now().strftime('%d %b %Y, %H:%M')}<br>
            <span style='color:#059669'>D+2: {d2.strftime('%d %b %Y')}</span>
        </div>
        """, unsafe_allow_html=True)

        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()

    if "Input" in page:
        page_input()
    elif "Consolidated" in page:
        page_consolidated()
    elif "Live" in page:
        page_live()
    else:
        page_dashboard()


if __name__ == "__main__":
    main()
